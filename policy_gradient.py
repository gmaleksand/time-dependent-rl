from jax import grad , jit, random, lax, value_and_grad
import jax.numpy as jnp
import numpy as np
from jax.example_libraries import optimizers, stax #from jax.experimental import optimizers, stax (use with the old version of JAX)
from jax. tree_util import tree_flatten
import time
from matplotlib import pyplot as plt
import os
from environments import *
import pickle
from openpyxl import load_workbook

#Initial parameters
N_epochs = 502
N_MC = 256
t = 40.
deltat = 0.2
#reward_equation="(-4*(x-np.sqrt(2)/2)**2 - p**2) - 10*(1-done_now)"    #double well
reward_equation="np.abs(x)" #cosh potential
#reward_equation = "- 2*(r-3*np.pi)**2"# - (pr**2)/2" #2D potential

env = fixed_env(N_MC=N_MC,
                reward_equation=reward_equation,   #Default reward equation
                t=t,
                deltat = deltat
)
K = 1
T0 = 1.
Tdecay = 400.0
seed = 0
begin_time = int(time.time())

#Initializing the neural network

rng = np.random.default_rng(seed)
jrng = random.PRNGKey(seed)

N_states = env.N_states
N_actions = env.N_actions
n_time_steps = env.n_time_steps-1
temperature = T0

#Hyperparameters of the neural network

step_size = 0.002
b1 = .9 
b2 = .999
eps = 1e-8
l2_param = 0.01
discount_rate = 1
layer_1 = 128
layer_2 = 128
layer_3 = 64

optimizer_name='adam'
if optimizer_name=='sgd':
    opt_init, opt_update, get_params = optimizers.sgd(step_size)
if optimizer_name=='adam':
    opt_init, opt_update, get_params = optimizers.adam(step_size, b1=b1, b2=b2, eps=eps)


initialize_params, predict = stax.serial(
                                            stax.Dense(layer_1), stax.Relu,
                                            stax.Dense(layer_2), stax.Relu,
                                            #stax.Dense(layer_3), stax.Sigmoid,
                                            #stax.Dense(64), stax.Relu,
                                            stax.Dense(N_actions), stax.LogSoftmax
                                        )
output_shape, initial_params = initialize_params(jrng, (-1,n_time_steps,N_states))

rewards = np.empty((N_MC,n_time_steps))

opt_state = opt_init(initial_params)
state = np.zeros(N_states)
statelist = np.zeros((N_MC, n_time_steps,N_states))
actionlist = np.zeros((N_MC, n_time_steps),dtype=int)
returns = np.zeros((N_MC, n_time_steps))
returns_mask = np.ones((N_MC, n_time_steps), dtype=bool)      #array which will contain masks
mean_return = np.zeros(N_epochs)
min_return = np.zeros(N_epochs)
max_return = np.zeros(N_epochs)
entropylist = np.zeros(N_epochs)             #array which will save the entropy of the policy
successful_trajectories = np.zeros(N_epochs) #array which will save the number of successful trajectories
Tlist = np.zeros(N_epochs)
Tlist = T0*np.exp(-np.arange(N_epochs)/Tdecay)
#Tlist[800:] = T0*np.exp(-800/Tdecay)

#Create a folder containing the results of the experiment
directory =  f'data/cosh_pseudovar_height&t={begin_time:d}'
os.mkdir(directory)

def details_to_table():# add all relevant parameters of the simulation in the table of experiments
    wb = load_workbook('data/details.xlsx')
    ws = wb.active
    col = ws.max_column+1
    ws.cell(row=1,column=col,value=directory)
    ws.cell(row=2,column=col,value='PG')            #Algorithm -> Policy gradient (PPO with K=1)
    #ws.cell(row=3,column=col,value='0')             #Common layers -> none
    ws.cell(row=4,column=col,value=N_MC) 
    ws.cell(row=5,column=col,value=t)
    ws.cell(row=6,column=col,value=deltat)
    """                                      NN architecture                                """
    ws.cell(row=8,column=col,value=f'{layer_1:d} Relu {layer_2:d} Relu {N_actions:d} LogSoftmax\n\n')
    #ws.cell(row=9,column=col,value='baselne')# no critic
    ws.cell(row=10,column=col,value=optimizer_name)
    ws.cell(row=11,column=col,value=step_size)
    ws.cell(row=12,column=col,value=b1)
    ws.cell(row=13,column=col,value=b2)
    ws.cell(row=14,column=col,value=eps)
    #ws.cell(row=15,column=col,value=cr_step_size)
    ws.cell(row=16,column=col,value=l2_param)
    #ws.cell(row=17,column=col, value=critic_param)
    ws.cell(row=18,column=col,value=discount_rate)
    ws.cell(row=19,column=col,value='cosh_var_height')
    ws.cell(row=20,column=col,value=seed)

    #ws.cell(row=22,column=col,value=env.usemasks)
    ws.cell(row=23,column=col,value=reward_equation)
    #ws.cell(row=24,column=col,value=K)
    ws.cell(row=25,column=col,value=T0)
    ws.cell(row=26,column=col,value=Tdecay)
    ws.cell(row=27,column=col,value="0.05 limit for success")       #Notes
    ws.cell(row=33,column=col,value=env.Fpower)
    wb.save('data/details.xlsx')
details_to_table()


# save own code
with open('policy_gradient.py') as f, open(directory + '/self.py', 'w') as d:
    d.write(f.read())
with open('environments_2D.py') as f, open(directory + '/env.py', 'w') as d:
    d.write(f.read())

def cost(params, oldparams, batch, temperature, eps=0.1):

    #states, actions, returns, returns_mask = batch      #added masks to batch
    states, actions, returns = batch

    preds = predict(params, states)
    oldpreds = predict(oldparams, states)

    #baseline = jnp.nan_to_num(jnp.true_divide(returns.sum(axis=0),returns_mask.sum(axis=0)))
    baseline = jnp.mean(returns,axis=0)

    preds_trajectory = jnp.take_along_axis(preds, jnp.expand_dims(actions, axis=2), axis=2).squeeze()
    """     PPO    """
    oldpreds_trajectory = lax.stop_gradient(jnp.take_along_axis(oldpreds, jnp.expand_dims(actions, axis=2), axis=2).squeeze())
    A = (returns-baseline)                                                                 #Advantage
    imp = jnp.exp(preds_trajectory-oldpreds_trajectory)                                    #Importance sampling ratio
    #c = -jnp.mean(jnp.sum(jnp.minimum(imp*A,jnp.clip(imp,1-eps,1+eps)*A), axis=1))         #Simple cost
    c = -jnp.mean(jnp.mean(jnp.minimum(imp*A,jnp.clip(imp,1-eps,1+eps)*A), axis=1)) -jnp.mean(temperature*entropy(preds)) + l2reg(params, l2_param)
    """     END PPO"""
    """     policy gradient"""
    #c = -jnp.mean(jnp.mean(preds_trajectory * lax.stop_gradient(returns_mask*(returns - baseline)), axis=1)) -jnp.mean(temperature*entropy(preds)) + l2reg(params, l2_param)     #masked pg
    #c = -jnp.mean(jnp.sum(preds_trajectory*(returns-baseline), axis=1)) + l2reg(params, l2_param)         #policy gradient
    return c

def entropy(preds):
    H = -jnp.sum(preds*preds, axis=2)
    return H


def true_entropy(preds):
    H = -jnp.sum(preds*jnp.exp(preds), axis=2)
    return H

def l2reg(params, l2_param):
    return l2_param*jnp.sum(jnp.array([jnp.sum(jnp.abs(theta)**2) for theta in tree_flatten(params)[0] ]))

@jit #speed-up
def update(i, opt_state, oldparams, batch, temperature):
    current_params = get_params(opt_state)
    grad_params = grad(cost)(current_params, oldparams, batch, temperature)
    return opt_update(i, grad_params, opt_state)


def render():                    #Training curve

    plt.rc('font', size=15)        #Resize font
    plt.figure(figsize=(19.2,10.8))
    episodes=list(range(N_epochs))
    plt.plot(episodes[:-1],mean_return[:-1],label='Mean return')
    #plt.fill_between(episodes[:-1],min_return[:-1],max_return[:-1], color='k', alpha=0.25)
    plt.xlabel('Epoch')
    plt.ylabel('Mean return')
    plt.grid(True)
    plt.title('Returns per epoch')
    #plt.legend()

    #Save plot
    file = '/train&epochs={:d}.png'.format(N_epochs)
    try:   
        plt.savefig(directory+file)
    except:
        plt.savefig(directory+'/train&epochs={:d}&exception_time={:d}.png'.format(N_epochs,int(time.time())))
    plt.close('all')
    #""" Successful trajectories plot
    plt.figure(figsize=(19.2,10.8))
    plt.plot(episodes,successful_trajectories,label='Successful trajectories')
    plt.xlabel('Epoch')
    plt.ylabel('Successful trajectories')
    file = '/successful_trajectories.png'
    plt.savefig(directory+file)
    plt.close('all')
    #"""
    ### Agent entropy curve
    plt.figure(figsize=(19.2,10.8))
    plt.plot(episodes,entropylist,label='Policy entropy')
    plt.xlabel('Epoch')
    plt.ylabel('Entropy')
    file = '/entropy.png'
    plt.savefig(directory+file)
    plt.close('all')


    ### Loss on which the agent is learning
    plt.figure(figsize=(19.2,10.8))
    plt.xlabel("Epoch")
    plt.ylabel("Learning loss")
    plt.plot(episodes[:-1],-2*entropylist[:-1]*Tlist[:-1]+mean_return[:-1]/n_time_steps)
    plt.plot(episodes[:-1],mean_return[:-1]/n_time_steps)
    file = '/loss.png'
    plt.savefig(directory+file)
    plt.close('all')

def save_total_state():
    paramsfile = open(directory+'/{:d}_total_backup.txt'.format(epoch), 'wb')
    pickle.dump([optimizers.unpack_optimizer_state(opt_state), 
        rng.bit_generator.state, 
        entropylist, 
        mean_return, 
        successful_trajectories,
        min_return,
        max_return,
        Tlist*n_time_steps], paramsfile)
    paramsfile.close()

#Vectorized random choice
#https://stackoverflow.com/questions/47722005/vectorizing-numpy-random-choice-for-given-2d-array-of-probabilities-along-an-a
def random_choice_prob_index(a, axis=1):
    r = np.expand_dims(rng.random(a.shape[1-axis]), axis=axis)
    return (a.cumsum(axis=axis) > r).argmax(axis=axis)

if False:    #Use the pretrained strategy
    paramsfile = open('data/2D_var_min&t=1673162824/4000_total_backup.txt','rb')
    ne = 4002
    x, y, entropylist[:ne], mean_return[:ne], successful_trajectories[:ne], min_return[:ne], max_return[:ne], _= pickle.load(paramsfile)
    opt_state = optimizers.pack_optimizer_state(x)
    rng.bit_generator.state = y
    paramsfile.close()

epoch = 0



# Start training
while epoch < N_epochs:

    start_time = time.time()

    if epoch%250 == 0:                          #Simulation backup
        save_total_state()

    nnparams = get_params(opt_state)        #Get the current policy in policy matrix

    #if epoch==N_epochs-1:   #Last test performed with grid-like initial state
    #    z0 = np.array([np.linspace(-1,1,num=N_MC),rng.uniform(low=-2,high=2,size=N_MC)])
    #    env.change_initial_state(z0, multiple_initial_states=True)


    env.reset()                             #Choose the first state
    dones = np.zeros(N_MC)
    timestep = 0
    while timestep<n_time_steps:
        #print(np.shape())
        p = np.exp(predict(nnparams, env.state))
        action = random_choice_prob_index(p)    #Vectorized random choice
        if epoch>=N_epochs-2: 
            action = np.argmax(p,axis=1)        #Greedy tests
        statelist[:,timestep,:], rewards[:,timestep], dones = env.step(action, dones)
        #statelist[:,timestep,:], rewards[:,timestep] = env.step(action)
        actionlist[:,timestep] = action

        #returns_mask[:,timestep] = np.invert(dones)    #mask


        timestep += 1

    #returns[:,:] = jnp.cumsum(rewards[:,::-1], axis=1)[:,::-1]                     #non-discounted cumulative sum of the rewards
    discounts = jnp.power(discount_rate,jnp.arange(n_time_steps))
    returns[:,:] = jnp.cumsum((rewards*discounts)[:,::-1], axis=1)[:,::-1]/discounts    #discounted cumulative sum (+3.5 s for 800 calls, N_MC = 512, 200 timesteps)
    #print(returns)

    batch = statelist, actionlist, returns#, returns_mask                     #added masks to batch

    oldparams = get_params(opt_state)
    """
    if epoch<800:
        temperature = T0*np.exp(-epoch/Tdecay)      #Used for entropy
    else:
        temperature = T0*np.exp(-800/Tdecay)
    """
    if epoch !=N_epochs-1:  #Do not learn in the test epoch
        for u in range(K):                          #K=1 for policy gradient
            opt_state = update(epoch*K+u, opt_state, oldparams, batch, Tlist[epoch])

    #The following calculations are needed for recording and display
    mean_return[epoch] = np.mean(returns[:,0])
    min_return[epoch] = np.min(returns[:,0])
    max_return[epoch] = np.max(returns[:,0])
        
    nnparams = get_params(opt_state)
    preds = predict(nnparams, statelist)
    entropylist[epoch] = jnp.mean(true_entropy(preds))
    

    #Print results of the epoch
    epoch_time = time.time()-start_time
    print("Epoch {} in {:0.2f} sec".format(epoch, epoch_time) )
    print("Returns: {:0.4f} mean; {:0.4f} min; {:0.4f} max".format(mean_return[epoch],min_return[epoch], max_return[epoch]) )
    print("Successful trajectories: {:d}/{:d}\n".format(np.sum(dones),N_MC))
    successful_trajectories[epoch] = np.sum(dones)

    if epoch%20 == 0 or epoch==N_epochs-1:
    #if epoch>=500 and epoch<=600:

        y = np.arange(-1, 1, 0.1)
        x = np.arange(-1.5, 1.5, 0.15)
        states = np.zeros((20,20,env.N_states))
        states[:,:,1], states[:,:,0] = np.meshgrid(x,y)

        nnparams = get_params(opt_state)
        probs = np.exp(predict(nnparams, states))
        if epoch!=N_epochs-1:
            env.Render((directory,epoch,rewards), probs)#, trajectory_numbers_to_show=range(N_MC))              #Phase diagram
        else:
            env.Render((directory,'greedy',rewards), probs, trajectory_numbers_to_show=range(N_MC))    #Output all trajectories in the final diagram
            env.save_phase_diagram((directory,'greedy',returns), probs)


    if (epoch-1)%50 == 0 or epoch==N_epochs-1:              #Print the training curve
        render()

    if epoch%250 == 0:

        if (epoch%20 != 0) and (epoch != N_epochs-1):
            y = np.arange(-1, 1, 0.1)
            x = np.arange(-1.5, 1.5, 0.15)
            states = np.zeros((20,20,env.N_states))
            states[:,:,1], states[:,:,0] = np.meshgrid(x,y)

            nnparams = get_params(opt_state)
            probs = np.exp(predict(nnparams, states))


        env.Render_best((directory,str(epoch)+'best',rewards), probs, show_arrows=False)
        env.save_phase_diagram((directory,epoch,returns), probs)

    if epoch == N_epochs - 1:                               #Output the training curve to file
        train_curve_file = open(directory+'/training_curve.txt', 'wb')
        pickle.dump(mean_return, train_curve_file)
        train_curve_file.close()
        print("total time: ", time.time()-begin_time)

        try:
            wb = load_workbook('data/details.xlsx')
            ws = wb.active
            col = ws.max_column
            ws.cell(row=28,column=col,value=mean_return[N_epochs-2])                #Final (train) mean return
            ws.cell(row=29,column=col,value=successful_trajectories[N_epochs-2])    #Final (train) successful trajectories
            ws.cell(row=30,column=col,value=mean_return[N_epochs-1])                #Test mean return
            ws.cell(row=31,column=col,value=successful_trajectories[N_epochs-1])    #Test successful trajectories
            wb.save('data/details.xlsx')
        except:
            print("Unsuccessful recording of the final performance")

    epoch += 1

print('\a')     #Alert when the program has finished