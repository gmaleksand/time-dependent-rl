from jax import grad , jit, random, lax, value_and_grad
import jax.numpy as jnp
import numpy as np
from jax.example_libraries import optimizers, stax #from jax.experimental import optimizers, stax (use with the old version of JAX)
from jax. tree_util import tree_flatten
import time
from matplotlib import pyplot as plt
import os
from environments import *
import pickle
from openpyxl import load_workbook

#Initial parameters
N_epochs = 4002
N_MC = 512
t = 40.
deltat = 0.2
reward_equation="(-4*(x-np.sqrt(2)/2)**2 - p**2) - 10*(1-done_now)"    #double well
#reward_equation="np.abs(x)" #cosh potential

env = var_env(N_MC=N_MC,
                reward_equation=reward_equation,   #Default reward equation
                t=t,
                deltat = deltat
)
K = 1
T0 = 1.
Tdecay = 400.0
seed = 0
begin_time = int(time.time())

#Initializing the neural network

rng = np.random.default_rng(seed)   #Numpy random number generator
jrng = random.PRNGKey(seed)         #JAX random number generator

N_states = env.N_states
N_actions = env.N_actions
n_time_steps = env.n_time_steps-1
temperature = T0

#Hyperparameters of the neural network

step_size = 0.0001
b1 = .9 
b2 = .999
eps = 1e-8
l2_param = 0.01
discount_rate = 1
layer_1 = 128
layer_2 = 128
layer_3 = 64

optimizer_name='adam'
if optimizer_name=='sgd':
    opt_init, opt_update, get_params = optimizers.sgd(step_size)
if optimizer_name=='adam':
    opt_init, opt_update, get_params = optimizers.adam(step_size, b1=b1, b2=b2, eps=eps)


initialize_params, predict = stax.serial(
                                            stax.Dense(layer_1), stax.Relu,
                                            stax.Dense(layer_2), stax.Relu,
                                            #stax.Dense(layer_3), stax.Sigmoid,
                                            #stax.Dense(64), stax.Relu,
                                            stax.Dense(N_actions), stax.LogSoftmax
                                        )
output_shape, initial_params = initialize_params(jrng, (-1,n_time_steps,N_states))

rewards = np.empty((N_MC,n_time_steps))

opt_state = opt_init(initial_params)
state = np.zeros(N_states)
statelist = np.zeros((N_MC, n_time_steps,N_states))
actionlist = np.zeros((N_MC, n_time_steps),dtype=int)
returns = np.zeros((N_MC, n_time_steps))
returns_mask = np.ones((N_MC, n_time_steps), dtype=bool)      #array which will contain masks
### Data recorded for further analysis
mean_return = np.zeros(N_epochs)
min_return = np.zeros(N_epochs)
max_return = np.zeros(N_epochs)
successful_trajectories = np.zeros(N_epochs) #array which will save the number of successful trajectories
entropylist = np.zeros(N_epochs)             #array which will save the entropy of the policy
z0_array = np.zeros((2,N_epochs))
Tlist = np.zeros(N_epochs)
Tlist = T0*np.exp(-np.arange(N_epochs)/Tdecay)

directory =  f'data/double_well_var_tdep_diff&t={begin_time:d}'


os.mkdir(directory)
def details_to_table():# add all relevant parameters of the simulation in the table of experiments
    wb = load_workbook('data/details.xlsx')
    ws = wb.active
    col = ws.max_column+1
    ws.cell(row=1,column=col,value=directory)
    ws.cell(row=2,column=col,value='PG')            #Algorithm -> Policy gradient (PPO with K=1)
    #ws.cell(row=3,column=col,value='0')             #Common layers -> none
    ws.cell(row=4,column=col,value=N_MC) 
    ws.cell(row=5,column=col,value=t)
    ws.cell(row=6,column=col,value=deltat)
    """                                      NN architecture                                """
    ws.cell(row=8,column=col,value=f'{layer_1:d} Relu {layer_2:d} Relu {N_actions:d} LogSoftmax\n\n')
    #ws.cell(row=9,column=col,value='baselne')# no critic
    ws.cell(row=10,column=col,value=optimizer_name)
    ws.cell(row=11,column=col,value=step_size)
    ws.cell(row=12,column=col,value=b1)
    ws.cell(row=13,column=col,value=b2)
    ws.cell(row=14,column=col,value=eps)
    #ws.cell(row=15,column=col,value=cr_step_size)
    ws.cell(row=16,column=col,value=l2_param)
    #ws.cell(row=17,column=col, value=critic_param)
    ws.cell(row=18,column=col,value=discount_rate)
    ws.cell(row=19,column=col,value='double_well_var_tdep_diff')
    ws.cell(row=20,column=col,value=seed)

    #ws.cell(row=22,column=col,value=env.usemasks)
    ws.cell(row=23,column=col,value=reward_equation)
    #ws.cell(row=24,column=col,value=K)
    ws.cell(row=25,column=col,value=T0)
    ws.cell(row=26,column=col,value=Tdecay)
    ws.cell(row=27,column=col,value="0.05 limit for success")       #Notes
    ws.cell(row=33,column=col,value=env.Fpower)
    wb.save('data/details.xlsx')
details_to_table()

# save own code
with open('pg_different_states.py') as f, open(directory + '/self.py', 'w') as d:
    d.write(f.read())
with open('environments.py') as f, open(directory + '/env.py', 'w') as d:
    d.write(f.read())


def cost(params, oldparams, batch, temperature, eps=0.1):

    states, actions, returns = batch

    preds = predict(params, states)
    oldpreds = predict(oldparams, states)

    baseline = jnp.mean(returns,axis=0)
    
    preds_trajectory = jnp.take_along_axis(preds, jnp.expand_dims(actions, axis=2), axis=2).squeeze()
    """     PPO  """
    oldpreds_trajectory = lax.stop_gradient(jnp.take_along_axis(oldpreds, jnp.expand_dims(actions, axis=2), axis=2).squeeze())
    A = (returns-baseline)                                                                 #Advantage
    imp = jnp.exp(preds_trajectory-oldpreds_trajectory)                                    #Importance sampling ratio
    #c = -jnp.mean(jnp.sum(jnp.minimum(imp*A,jnp.clip(imp,1-eps,1+eps)*A), axis=1))         #Simple cost
    c = -jnp.mean(jnp.mean(jnp.minimum(imp*A,jnp.clip(imp,1-eps,1+eps)*A), axis=1)) -jnp.mean(temperature*entropy(preds)) + l2reg(params, l2_param)
    """ policy gradient """
    #c = -jnp.mean(jnp.sum(preds_trajectory * lax.stop_gradient(returns_mask*(returns - baseline)), axis=1)) + l2reg(params, l2_param)     #masked pg
    #c = -jnp.mean(temperature*entropy(preds)) -jnp.mean(jnp.mean(preds_trajectory * lax.stop_gradient(returns_mask*(returns - baseline)), axis=1)) + l2reg(params, l2_param)     #masked pg
    return c

def entropy(preds):
    H = -jnp.sum(preds*preds, axis=2)
    return H


def true_entropy(preds):
    H = -jnp.sum(preds*jnp.exp(preds), axis=2)
    return H

def l2reg(params, l2_param):
    return l2_param*jnp.sum(jnp.array([jnp.sum(jnp.abs(theta)**2) for theta in tree_flatten(params)[0] ]))

@jit #speed-up
def update(i, opt_state, oldparams, batch, temperature):
    current_params = get_params(opt_state)
    grad_params = grad(cost)(current_params, oldparams, batch, temperature)  #cost_value, grad_params = value_and_grad(cost)(current_params, oldparams, batch)
    return opt_update(i, grad_params, opt_state)


def render():                    #Training curve

    plt.rc('font', size=15)        #Resize font
    plt.figure(figsize=(19.2,10.8))
    episodes=list(range(N_epochs))
    plt.plot(episodes[:-1],mean_return[:-1],'o',label='Mean return')
    plt.fill_between(episodes[:-1],min_return[:-1],max_return[:-1], color='k', alpha=0.25)
    plt.xlabel('Epoch')
    plt.ylabel('Mean return')
    plt.grid(True)
    plt.title('Returns per epoch')
    #plt.legend()

    #Save plot
    file = '/train&epochs={:d}.png'.format(N_epochs)
    try:   
        plt.savefig(directory+file)
    except:
        plt.savefig(directory+'/train&epochs={:d}&exception_time={:d}.png'.format(N_epochs,int(time.time())))
    plt.close('all')
    """ Successful trajectories plot
    plt.figure(figsize=(19.2,10.8))
    plt.plot(episodes,successful_trajectories,label='Successful trajectories')
    plt.xlabel('Epoch')
    plt.ylabel('Successful trajectories')
    file = '/successful_trajectories.png'
    plt.savefig(directory+file)
    plt.close('all')
    """
    ### Agent entropy curve
    plt.figure(figsize=(19.2,10.8))
    plt.plot(episodes,entropylist,label='Policy entropy')
    plt.xlabel('Epoch')
    plt.ylabel('Entropy')
    file = '/entropy.png'
    plt.savefig(directory+file)
    plt.close('all')

def save_total_state():
    paramsfile = open(directory+'/{:d}_total_backup.txt'.format(epoch), 'wb')
    pickle.dump([optimizers.unpack_optimizer_state(opt_state), 
        rng.bit_generator.state, 
        entropylist, 
        mean_return, 
        successful_trajectories,
        min_return,
        max_return,
        Tlist*n_time_steps], paramsfile)
    paramsfile.close()

def load_total_state(filename):
    paramsfile = open(filename,'rb')
    nnparams, x, entropylist, mean_return, successful_trajectories = pickle.load(paramsfile)
    rng.bit_generator.state = x
    paramsfile.close()
    return nnparams, entropylist, mean_return, successful_trajectories


#Vectorized random choice
#https://stackoverflow.com/questions/47722005/vectorizing-numpy-random-choice-for-given-2d-array-of-probabilities-along-an-a
def random_choice_prob_index(a, axis=1):
    r = np.expand_dims(rng.random(a.shape[1-axis]), axis=axis)
    return (a.cumsum(axis=axis) > r).argmax(axis=axis)


#nnparams, entropylist[:4002], mean_return[:4002], successful_trajectories[:4002] = load_total_state('Trainfigures/z0=randomPPO&t=1662215729/4000_total_backup.txt')     # Restore previous state
#opt_state = opt_init(nnparams)

""" Use the pretrained strategy
paramsfile = open('data/double_well_var_tdep_diff&t=1668782230/2250_total_backup.txt','rb')
x, y, entropylist[:4002], mean_return[:4002], successful_trajectories[:4002] = pickle.load(paramsfile)
opt_state = optimizers.pack_optimizer_state(x)
rng.bit_generator.state = y
paramsfile.close()
#"""
epoch = 0

# Start training
#for epoch in range(N_epochs):
while epoch < N_epochs:
        
    if epoch%250 == 0:                          #Simulation backup
        save_total_state()
    start_time = time.time()

    if epoch==N_epochs-1:   #Last test performed with grid-like initial state
        z0 = np.array([np.linspace(-1,1,num=N_MC),rng.uniform(low=-2,high=2,size=N_MC)])
        env.change_initial_state(z0, multiple_initial_states=True)
        z0_array[:,epoch] = [0,0]
    elif epoch%10==0:        #Random initial state each 10-th epoch
    #elif successful_trajectories[epoch-1]>1:                                               #Switch the initial state if 1 trajectory is learned
        p0 = rng.uniform(low=-1.,high=1.)
        #p0 = (rng.random()*4.+1.)*rng.choice([-1,1])
        x0 = rng.uniform(low=-2.,high=2.)
        z0 = [p0,x0]
        env.change_initial_state(z0)
        z0_array[:,epoch] = z0
    #elif epoch%2==0:
    #    z0 = -env.z0[:,0]
    #    env.change_initial_state(z0)
    #    z0_array[:,epoch] = z0
    else:
        env.reset()
        z0_array[:,epoch] = z0
    #"""

    nnparams = get_params(opt_state)        #Get the current policy in policy matrix

    dones = np.zeros(N_MC)
    timestep = 0
    while timestep<n_time_steps:

        p = np.exp(predict(nnparams, env.state))
        action = random_choice_prob_index(p)    #Vectorized random choice
        if epoch>=N_epochs-2:
            action = np.argmax(p,axis=1)        #Tests should be greedy
        statelist[:,timestep,:], rewards[:,timestep], dones = env.step(action, dones)
        actionlist[:,timestep] = action

        #returns_mask[:,timestep] = np.invert(dones)    #mask


        timestep += 1

    #returns[:,:] = jnp.cumsum(rewards[:,::-1], axis=1)[:,::-1]                     #non-discounted cumulative sum of the rewards
    discounts = jnp.power(discount_rate,jnp.arange(n_time_steps))
    returns[:,:] = jnp.cumsum((rewards*discounts)[:,::-1], axis=1)[:,::-1]/discounts    #discounted cumulative sum (+3.5 s for 800 calls, N_MC = 512, 200 timesteps)

    batch = statelist, actionlist, returns#, returns_mask                     #added masks to batch

    oldparams = get_params(opt_state)
    if epoch<400:
        temperature = T0*np.exp(-epoch/Tdecay)      #Used for entropy
    else:
        temperature = T0*np.exp(-400/Tdecay)        #Entropy doesn't decay

    if epoch < N_epochs-2:  #The network must not be learning during the test epochs.
        for u in range(K):                          #K=1 for policy gradient
            opt_state = update(epoch*K+u, opt_state, oldparams, batch, temperature)

    mean_return[epoch] = np.mean(returns[:,0])
    min_return[epoch] = np.min(returns[:,0])
    max_return[epoch] = np.max(returns[:,0])

    nnparams = get_params(opt_state)
    preds = predict(nnparams, statelist)
    entropylist[epoch] = jnp.mean(true_entropy(preds))


    #Print results of the epoch
    epoch_time = time.time()-start_time
    print("Epoch {} in {:0.2f} sec".format(epoch, epoch_time) )
    print("Returns: {:0.4f} mean; {:0.4f} min; {:0.4f} max".format(mean_return[epoch],min_return[epoch], max_return[epoch]) )

    #print("Successful trajectories: {:d}/{:d}\n".format(np.sum(dones),N_MC))
    successful_trajectories[epoch] = np.sum(dones)

    if epoch%20 == 0 or epoch==N_epochs-1:

        y = np.arange(-1, 1, 0.1)
        x = np.arange(-1.5, 1.5, 0.15)
        states = np.empty((20,20,env.N_states))
        states[:,:,1], states[:,:,0] = np.meshgrid(x,y)

        nnparams = get_params(opt_state)
        probs = np.exp(predict(nnparams, states))
        if epoch!=N_epochs-1:
            env.Render((directory,epoch,returns), probs)#, trajectory_numbers_to_show=range(N_MC))              #Phase diagram
        else:
            env.Render((directory,'greedy',returns), probs, trajectory_numbers_to_show=range(N_MC))    #Output all trajectories in the final diagram
            env.save_phase_diagram((directory,'greedy',returns), probs)

    if epoch%250 == 0:
        env.save_phase_diagram((directory,epoch,returns), probs)



    if (epoch-1)%50 == 0 or epoch==N_epochs-1:              #Print the training curve
        render()
    if epoch == N_epochs - 1:                               #Output the training curve to file
        train_curve_file = open(directory+'/training_curve.txt', 'wb')
        pickle.dump(mean_return, train_curve_file)
        train_curve_file.close()

        with open(directory+'/z0.txt','w') as z0_file:
            for i in range(N_epochs):
                z0_file.write('[{:.1f}, {:.1f}]; ({}/{:d})\n'.format(z0_array[0,i],z0_array[1,i],successful_trajectories[i],N_MC))

        try:
            wb = load_workbook('data/details.xlsx')
            ws = wb.active
            col = ws.max_column
            ws.cell(row=28,column=col,value=mean_return[N_epochs-2])                #Final (train) mean return
            ws.cell(row=29,column=col,value=successful_trajectories[N_epochs-2])    #Final (train) successful trajectories
            ws.cell(row=30,column=col,value=mean_return[N_epochs-1])                #Test mean return
            ws.cell(row=31,column=col,value=successful_trajectories[N_epochs-1])    #Test successful trajectories
            wb.save('data/details.xlsx')
        except:
            print("Unsuccessful recording of the final performance")

        print("total time: ", time.time()-begin_time)
    epoch += 1

print('\a')     #Alert when the program has finished